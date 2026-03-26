import json
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from tabulate import tabulate



class Database:
    def __init__(self):
        self.df = None
        self.caminho = os.path.join(os.path.dirname(__file__), "data.json")

    def carregar_dados(self):
        if not os.path.exists(self.caminho):
            return []
        
        with open(self.caminho, "r", encoding="utf-8") as f:
            try:
                return json.load(f)
            except json.JSONDecodeError:
                return []

    def salvar_dados(self, dados):
        with open(self.caminho, "w", encoding="utf-8") as f:
            json.dump(dados, f, indent=4, ensure_ascii=False)

    def criar_dataframe(self):
        dados = self.carregar_dados()
        self.df = pd.DataFrame(dados)
        return self.df

    def ler_dataframe(self):
        df = self.criar_dataframe()
        print(tabulate(df, headers='keys', tablefmt='psql'))

    def exportar_excel(self, caminho_excel):
        self.criar_dataframe()

        self.df.to_excel(caminho_excel, engine="openpyxl", index=False)
        return self.df
        
    def formatar_excel(self, caminho_excel):
        wb = load_workbook(caminho_excel)
        ws = wb.active

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3c3c3c", end_color="3c3c3c", fill_type='solid')
        center_aligment = Alignment(horizontal='center', vertical='bottom')
        cell_fill = PatternFill(start_color='BDBDBD', end_color='BDBDBD', fill_type='solid')

        for cell in ws[1]:
           cell.font = header_font
           cell.fill = header_fill
           cell.alignment = center_aligment
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.fill = cell_fill
                    cell.alignment = center_aligment

        full_range = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
        ws.auto_filter.ref = full_range

        wb.save(caminho_excel)
